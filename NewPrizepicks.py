
import json
import pandas as pd
import time
import datetime
import pytz
from playwright.sync_api import sync_playwright
from playwright_stealth import stealth_sync
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from google.oauth2.credentials import Credentials
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

# === CONFIG ===
LEAGUES_TO_FETCH = [1, 2, 3, 5, 82]  # PGA, MLB, WNBA, TENNIS, SOCCER
UPLOAD_INTERVAL_MINUTES = 60
GOOGLE_SHEET_FILE_ID = "1Ep3e2ekNYhcYok1bamtwzNUhIA1h3grIttTuhlgkPuw"
ENABLE_FETCH = True
mountain_tz = pytz.timezone("America/Phoenix")

LEAGUE_MAP = {
    1: "PGA", 2: "MLB", 3: "WNBA", 4: "NASCAR", 5: "TENNIS", 7: "NBA", 8: "NHL", 11: "CFL",
    12: "MMA", 15: "CFB", 20: "CBB", 25: "NFL2H", 35: "NFL1H", 42: "BOXING", 44: "NFLP",
    80: "NBA2H", 82: "SOCCER", 118: "COMBINE", 121: "LoL", 125: "F1", 131: "EUROGOLF",
    135: "KBO", 137: "XFINITY", 138: "TRUCKS", 145: "COD", 149: "NBA4Q", 150: "CFB2H",
    151: "EURO", 152: "NFL4Q", 153: "CFB4Q", 155: "CBB2H", 158: "GLEAGUE", 159: "VAL",
    161: "RL", 162: "CRICKET", 163: "NFLSZN", 165: "AFL", 172: "CFBSZN", 173: "NBASZN",
    174: "Dota2", 176: "WCBB", 184: "CFB1H", 188: "NBASZN2", 189: "NCAAB", 190: "MLBSZN",
    192: "NBA1Q", 193: "WNBA1H", 194: "WNBA2H", 195: "WNBQ", 226: "NHL3P", 227: "NHL1P",
    228: "LIVGOLF", 230: "LAX", 231: "MLBLIVE", 234: "NHL2P", 236: "NHLSZN", 238: "AUSNBL",
    242: "SOCCER1H", 243: "SOCCER2H", 244: "TENNIS LIVE", 245: "NFL1Q", 250: "NBA SERIES",
    251: "NHL SERIES", 252: "WNBASZN", 253: "FPA", 254: "LBSA", 255: "SACB", 256: "LPGA",
    260: "Power Slap", 261: "MLBSZN2", 262: "SOCCERSZN", 265: "CS2", 266: "SPECIALS",
    267: "HALO", 268: "APEX", 269: "DARTS", 270: "INDYCAR", 271: "UFL", 272: "WCBB2H",
    273: "PWHL", 274: "R6", 275: "WCWS", 284: "HANDBALL", 285: "BAD", 286: "TT",
    287: "EUROCUP", 288: "UNR", 289: "TGL", 290: "CBB1H", 291: "NASCARSZN", 298: "NPB",
    299: "BBL", 301: "KBL", 302: "BCL", 303: "Moments", 304: "NFL Draft", 305: "CLAX",
    306: "LAXSZN", 307: "TNC", 308: "WNBA1Q"
}

def fetch_prizepicks_arena(league_id, game_mode="arena"):
    url = f"https://api.prizepicks.com/projections?league_id={league_id}&per_page=250&single_stat=true&game_mode={game_mode}"
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)  # ✅ HEADLESS OFF
        context = browser.new_context()
        page = context.new_page()
        stealth_sync(page)
        for attempt in range(5):
            try:
                response = page.goto(url, timeout=20000)
                status = response.status if response else None
                if status == 200:
                    data = json.loads(response.text())
                    browser.close()
                    return data
                print(f"⚠️ Attempt {attempt + 1}: Bad status {status}")
            except Exception as e:
                print(f"❌ Attempt {attempt + 1} error: {e}")
            time.sleep(10)
        browser.close()
        raise Exception("❌ Failed to fetch data after retries.")

def process_and_export_to_writer(data, league_id, writer):
    # ✅ Use canonical player_id
    players = {
        p["id"]: {
            **p["attributes"],
            "player_id": p["id"]
        }
        for p in data.get("included", []) if p.get("type") == "new_player"
    }

    rows = []
    for proj in data.get("data", []):
        attr = proj.get("attributes", {})
        rel = proj.get("relationships", {})
        pid = rel.get("new_player", {}).get("data", {}).get("id")
        player = players.get(pid, {})
        rows.append({
            "player_id": player.get("player_id"),
            "Player": player.get("display_name"),
            "Stat": attr.get("stat_display_name"),
            "Line": attr.get("line_score"),
            "Odds Type": attr.get("odds_type"),
            "Description": attr.get("description"),
            "Start Time": attr.get("start_time"),
            "Status": attr.get("status"),
            "Player Image URL": player.get("image_url"),
            "Position": player.get("position"),
            "Team": player.get("team"),
        })

    df = pd.DataFrame(rows)
    sheet = LEAGUE_MAP.get(league_id, f"League_{league_id}")[:31]
    df.to_excel(writer, sheet_name=sheet, index=False)
    print(f"📦 {sheet} — {len(df)} props written.")

def apply_freeze_and_filters(filepath):
    wb = load_workbook(filepath)
    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            sheet.column_dimensions[col_letter].width = max_length + 2
    wb.save(filepath)

def overwrite_google_sheet(local_file, file_id):
    print("⏫ Uploading to Google Sheet...")
    try:
        creds = Credentials.from_authorized_user_file("token.json")
        service = build("drive", "v3", credentials=creds)
        media = MediaFileUpload(local_file, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resumable=True)
        file_metadata = {'mimeType': 'application/vnd.google-apps.spreadsheet'}
        updated_file = service.files().update(fileId=file_id, media_body=media, body=file_metadata, fields='id').execute()
        print(f"✅ Uploaded: https://docs.google.com/spreadsheets/d/{updated_file['id']}/edit")
    except Exception as e:
        print(f"❌ Upload error: {e}")

# === Main Loop ===
while True:
    print("\n🔄 Running PrizePicks Arena sync...")

    if ENABLE_FETCH:
        writer = pd.ExcelWriter("PrizePicks_Arena_AllLeagues.xlsx", engine="openpyxl")
        for lid in LEAGUES_TO_FETCH:
            print(f"\n🌐 Fetching League {lid} ({LEAGUE_MAP.get(lid, 'Unknown')})")
            try:
                data = fetch_prizepicks_arena(lid)
                process_and_export_to_writer(data, lid, writer)
            except Exception as e:
                print(str(e))
            print("⏳ Waiting 1 sec...")
            time.sleep(1)
        writer.close()
        print("✅ Exported all leagues.")

        # Add timestamp to cell M1
        wb = load_workbook("PrizePicks_Arena_AllLeagues.xlsx")
        ws = wb.active
        ws["M1"] = datetime.now(mountain_tz).strftime("%Y-%m-%d %I:%M:%S %p Arizona Time")
        wb.save("PrizePicks_Arena_AllLeagues.xlsx")

        apply_freeze_and_filters("PrizePicks_Arena_AllLeagues.xlsx")

    overwrite_google_sheet("PrizePicks_Arena_AllLeagues.xlsx", GOOGLE_SHEET_FILE_ID)
    print(f"⏱ Sleeping {UPLOAD_INTERVAL_MINUTES} min...")
    time.sleep(UPLOAD_INTERVAL_MINUTES * 60)
